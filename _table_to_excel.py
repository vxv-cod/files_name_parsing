import io # Импортируем модуль для работы со вводом/выводом байтов в памяти (BytesIO)
from typing import Any # Используем для указания типов (хотя явно не используется везде, полезно для аннотаций)
from openpyxl import Workbook # Импортируем класс для создания новой книги Excel
from openpyxl.worksheet.table import Table, TableStyleInfo # Импортируем классы для работы с табличным форматированием
from openpyxl.utils.cell import get_column_letter # Функция для получения буквы столбца по индексу (например, 1 -> 'A')
from openpyxl.styles import Alignment # Класс для настройки выравнивания ячеек


from rich import print # Импортируем функцию print из библиотеки rich для цветного/стилизованного вывода в консоль


def convert_bool_in_int(val):
    # Проверяем, если значение равно логическому True
    if val == True: 
        val = 1  # Преобразуем его в число 1
    # Проверяем, если значение равно логическому False
    if val == False: 
        val = 0  # Преобразуем его в число 0
    # Если входное значение является списком (например, список значений), преобразуем его в строку для записи в ячейку
    if isinstance(val, list): 
        val = str(val)
    # Возвращаем обработанное значение
    return val


def db_xlxs(filename: str, directory: str, cols: list[str], rows: list[dict], index_fist_row: int = 1): 
    wb = Workbook() # Создаем новый экземпляр книги Excel
    ws = wb.active # Получаем активный рабочий лист по умолчанию
    # Переименовать название листа с именем файла
    ws.title = filename 
    
    # Проверяем, если первая запись в 'rows' — это словарь (т.е., мы работаем со структурированными данными)
    if isinstance(rows[0], dict):
        # Преобразуем весь список словарей: извлекаем значения по именам колонок и преобразуем булевы типы в int
        rows: list = [[ convert_bool_in_int(row[col["name"]]) for col in cols] for row in rows ] 
        # Собираем имена столбцов (те, которые будут использоваться как заголовки)
        # names = [col["name"] for col in cols]
        # Собираем метки/описания для колонок
        labels = [col["label"] for col in cols]
    
    # Добавляем в лист метки (Labels) в первую строку данных
    ws.append(labels) 
    # Добавляем в лист имена столбцов (Names) во вторую строку данных
    # ws.append(names) 
    
    # Цикл для добавления реальных данных: добавляем каждую обработанную строку в лист
    for row in rows:
        ws.append(row)


    if index_fist_row < 1: # Проверка, чтобы начальная строка была не меньше 1
        index_fist_row = 1 # Если меньше 1, устанавливаем минимально допустимое значение 1


    # --- Расчет диапазонов для форматирования таблицы ---

    letter_fist_col = get_column_letter(1) # Получаем букву первой колонки (например, 'A')
    # Определяем последнюю букву столбца на основе количества колонок в данных
    letter_max_col = get_column_letter(len(rows[0])) 
    # Вычисляем максимальный номер строки: начальная строка + количество строк данных
    index_max_row = index_fist_row + len(rows) 
    # Формируем диапазон адресов ячеек (например, "A4:D15")
    ref = f"{letter_fist_col}{index_fist_row}:{letter_max_col}{index_max_row}"
    print(f"Вычисленный диапазон данных: {ref = }") # Выводим вычисленный диапазон для отладки
    
    # Вставляем пустые строки в начале (это часто требуется перед добавлением табличного стиля)
    # ws.insert_rows(1) 
    # ws.insert_rows(3)


    # --- Создание и применение стилей таблицы ---

    if not ws.tables.get('Table1'): # Проверяем, существует ли уже таблица с именем 'Table1' на листе
        # Создаем объект таблицы, указывая ему имя и рассчитанный диапазон ссылок (ref)
        table: Table = Table(displayName="Table1", ref=ref) 
        # Определяем стиль для этой таблицы (Medium2, скрываем первую/последнюю колонки, включаем чередование цветов строк)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                 showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        # Применяем определенный стиль к объекту таблицы
        table.tableStyleInfo = style 
        # Добавляем созданную таблицу на рабочий лист
        ws.add_table(table)
    
    # --- Автоматическая ширина столбцов и выравнивание текста ---
    """Рассчитывает и устанавливает ширину столбца на основе самой длинной ячейки."""
    
    # Получаем максимальный номер строки, где есть данные
    max_row = ws.max_row
    
    for col_idx, column in enumerate(ws.columns, 1):
        letter = get_column_letter(col_idx)
        
        max_length = 0
        
        # Итерируемся по всем строкам в этом столбце (от 1 до max_row)
        for row in range(1, max_row + 1):
            cell = ws[f'{letter}{row}']
            
            # Длина текста минус пробелы/символы форматирования (упрощенно)
            current_length = len(str(cell.value))
            
            if current_length > max_length:
                max_length = current_length

        # Устанавливаем ширину столбца, используя найденную максимальную длину.
        # Обычно в Excel 1 символ текста занимает примерно 1 единицу ширины.
        ws.column_dimensions[letter].width = max_length + 2 # +2 для запаса
        ws[f'{letter}1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        for idx, _ in enumerate(rows, 2):
            ws[f'{letter}{idx}'].alignment = Alignment(wrap_text=False, horizontal='left', vertical='center')
        
            
    # --- Сохранение результата ---

    # Формируем полный путь для сохранения файла в указанной директории
    new_filename = f"{directory}\\{filename}.xlsx" 
    # Сохраняем всю книгу на диск по новому пути
    wb.save(filename=new_filename) # Возвращаем None (или ничего) после успешного сохранения
        
    return True