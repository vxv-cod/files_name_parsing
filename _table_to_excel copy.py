import io
from typing import Any
# from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment

from rich import print



def convert_bool_in_int(val):
    # print(val, type(val))
    if val == True: val = 1
    if val == False: val = 0
    if isinstance(val, list): val = str(val)
    return val


def db_xlxs(filename: str, directory: str, cols: list[str], rows: list[dict], index_fist_row: int = 4):
    wb = Workbook()
    ws = wb.active
    # Переименовать название листа
    ws.title = filename
    if isinstance(rows[0], dict):
        rows: list = [[ convert_bool_in_int(row[col["name"]]) for col in cols] for row in rows ]
        names = [col["name"] for col in cols]
        labels = [col["label"] for col in cols]
    # добавим заголовки столбцов. Это должны быть строки
    ws.append(labels)
    ws.append(names)
    # Добавление строк
    for row in rows:
        ws.append(row)

    if index_fist_row < 1: index_fist_row = 1

    letter_fist_col = get_column_letter(1)
    letter_max_col = get_column_letter(len(rows[0]))
    index_max_row = index_fist_row + len(rows)
    ref = f"{letter_fist_col}{index_fist_row}:{letter_max_col}{index_max_row}"
    print(f"{ref = }")
    ws.insert_rows(1)
    ws.insert_rows(3)


    # создаем объект таблицы
    if not ws.tables.get('Table1'):
        table: Table = Table(displayName="Table1", ref=ref)
        # добавим стиль по умолчанию.
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        # # добавляем таблицу
        ws.add_table(table)
    
    # Автоширина колонок
    for idx, col in enumerate(ws.columns, 1):
        liter = get_column_letter(idx)
        ws.column_dimensions[liter].auto_size = True
        ws[f'{liter}2'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Сохранение полченных файлов
    new_filename = f"{directory}\\{filename}.xlsx"
    wb.save(filename=new_filename)
        
    return  
    # # '''Создание буфера обмена'''
    buffer = io.BytesIO()
    # '''Сохранение документа в буфер обмена'''
    wb.save(buffer)
    buffer.seek(0)
    # '''Способ через поток'''
    return StreamingResponse(content=buffer)


    '''Способ сохранения в темп-файл'''
    # headers = {
    #     'Content-Disposition': 'attachment', 
    #     'filename': f"{filename}.xlsx".encode("utf-8").decode("latin-1"),
    #     "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # }
    # import tempfile
    # with tempfile.NamedTemporaryFile(mode="w+b", suffix=".xlsx", delete=False) as temp_file:
    #     temp_file.write(buffer.getvalue())
    # return FileResponse(temp_file.name)

